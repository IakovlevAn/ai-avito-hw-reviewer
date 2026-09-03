from __future__ import annotations

from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

from app.models import Submission


def build_submission_workbook(submission: Submission) -> BytesIO:
    workbook = Workbook()
    summary = workbook.active
    summary.title = "Результат"
    summary.append(["Поле", "Значение"])
    summary.append(["Работа", submission.title])
    summary.append(["GitHub", submission.repository_url])
    summary.append(["Версия", submission.commit_sha or "—"])
    summary.append(["Ревьюер", submission.reviewer.name if submission.reviewer else "—"])
    summary.append(["Статус", submission.status])
    summary.append(["Баллы", sum(item.final_points or 0 for item in submission.criteria)])
    summary.append(["Максимум", submission.max_points])

    details = workbook.create_sheet("Критерии")
    details.append(
        [
            "Раздел",
            "Критерий",
            "Максимум",
            "Предложено",
            "Подтверждено",
            "Уверенность",
            "Комментарий",
            "Подтверждения",
        ]
    )
    for item in submission.criteria:
        details.append(
            [
                item.section,
                item.title,
                item.max_points,
                item.suggested_points,
                item.final_points,
                item.confidence,
                item.final_feedback or item.feedback,
                item.evidence_json,
            ]
        )

    header_fill = PatternFill("solid", fgColor="1F2023")
    for sheet in (summary, details):
        for cell in sheet[1]:
            cell.fill = header_fill
            cell.font = Font(color="FFFFFF", bold=True)
            cell.alignment = Alignment(vertical="center")
        sheet.freeze_panes = "A2"
        for column in sheet.columns:
            width = min(max(len(str(cell.value or "")) for cell in column) + 2, 60)
            sheet.column_dimensions[column[0].column_letter].width = width
        for row in sheet.iter_rows():
            for cell in row:
                cell.alignment = Alignment(vertical="top", wrap_text=True)

    buffer = BytesIO()
    workbook.save(buffer)
    buffer.seek(0)
    return buffer
